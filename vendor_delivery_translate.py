#!/usr/local/bin/python

import os
import csv
import sys
import re
from openpyxl import load_workbook
import shutil
# import PyQt5
# from PyQt5 import QtWidgets

# currently a one-off, only set up to work for ROMEO.

headers = ['Version Name',
           'Link',
           'Task',
           'Type',
           'Submitted For',
           'Description',
           'First Frame Text',
           'Last Frame Text',
           'Frame Count Text']

dirpath = ''

if len(sys.argv) < 2:
    print('Usage: vendor_delivery_translate.py /path/to/delivery/folder')
    exit()
dirpath = sys.argv[1]

if not os.path.exists(dirpath):
    print('Usage: vendor_delivery_translate.py /path/to/delivery/folder')
    exit()

if not os.path.isdir(dirpath):
    print('Usage: vendor_delivery_translate.py /path/to/delivery/folder')
    exit()

extra_files_dir = os.path.join(dirpath, 'support_files')
lut_files_dir = os.path.join(dirpath, 'luts')

shot_regexp_txt = r'([0-9]{3}_[A-Z]{3}_[0-9]{4})_'
sequence_regexp_txt = r'([A-Z]{3})_'
imgseq_regexp_txt = r'\.([0-9]+)\.'
version_name_regexp_text = r'([0-9]{3}_[A-Z]{3}_[0-9]{4}_[A-Za-z0-9\-_]+_v[0-9]+)'
version_name_wframe_regexp_text = r'([0-9]{3}_[A-Z]{3}_[0-9]{4})_([A-Za-z0-9\-_]+)_(v[0-9]+)\.([0-9]+)'
version_name_hack_regexp_text = r'(frame([0-9]+)_)'
version_regexp_text = r'_v([0-9]+)'

shot_regexp = re.compile(shot_regexp_txt)
sequence_regexp = re.compile(sequence_regexp_txt)
imgseq_regexp = re.compile(imgseq_regexp_txt)
version_name_regexp = re.compile(version_name_regexp_text)
version_regexp = re.compile(version_regexp_text)
version_name_wframe_regexp = re.compile(version_name_wframe_regexp_text)
version_name_hack_regexp = re.compile(version_name_hack_regexp_text)

filename_extras_regexp_text_list = [r'(.*)_vfx', r'(.*)_avid', r'(.*)_prores', r'(.*)_pr422']
filename_extras_regexp_list = [re.compile(pattern) for pattern in filename_extras_regexp_text_list]
master_files_list = []
master_files_ext_dict = {}
subform_header_columns = []
subform_rows = []

def extract_excel(filepath):
    print('Info: Parsing XLSX file: %s'%filepath)
    wb = load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    for header in ws.rows.next():
        subform_header_columns.append(header.value)

    for row in ws.iter_rows(min_row=2):
        rowdict = {}
        for idx, column in enumerate(row):
            rowdict[subform_header_columns[idx]] = column.value
        subform_rows.append(rowdict)
    old_filepath = os.path.join(extra_files_dir, os.path.basename(filepath))
    print('Info: Saving out XLSX file as: %s'%old_filepath)
    shutil.move(filepath, old_filepath)


def extract_csv(filepath):
    print('Info: Parsing CSV file: %s'%filepath)
    csvhandle = open(filepath, 'rU')
    csvreader = csv.reader(csvhandle)
    for header in csvreader.next():
        subform_header_columns.append(header)
    for row in csvreader:
        rowdict = {}
        for idx, column in enumerate(row):
            rowdict[subform_header_columns[idx]] = column
        subform_rows.append(rowdict)
    csvhandle.close()
    old_filepath = os.path.join(extra_files_dir, os.path.basename(filepath))
    print('Info: Saving out CSV file as: %s'%old_filepath)
    shutil.move(filepath, old_filepath)


def check_file_naming(filepath):
    # deal with tiff/tif
    (file_noext, file_dotext) = os.path.splitext(filepath)
    if file_dotext == '.tiff':
        new_filename_withext = '%s.tif'%file_noext
        print('Info: %s will be renamed to %s.' % (filepath, new_filename_withext))
        os.rename(filepath, new_filename_withext)
        filepath = new_filename_withext
    elif file_dotext == '.jpeg':
        new_filename_withext = '%s.jpg'%file_noext
        print('Info: %s will be renamed to %s.' % (filepath, new_filename_withext))
        os.rename(filepath, new_filename_withext)
        filepath = new_filename_withext

    filename = os.path.basename(filepath)
    b_renamed = False
    # remove extra os files
    if filename == '.DS_Store':
        os.unlink(filepath)
        return
    if filename == 'Thumbs.db':
        os.unlink(filepath)
        return
    if filename.find('_export') != -1:
        os.unlink(filepath)
        return
    filedir = os.path.dirname(filepath)
    if filedir.find('support_files') != -1:
        return
    filename_array = filename.split('.')
    filebase = filename_array[0]
    if len(filebase) < 1:
        return
    fileext = filename_array[-1]
    if fileext == 'xlsx':
        extract_excel(filepath)
        return
    elif fileext == 'csv':
        extract_csv(filepath)
        return
    else:
        if fileext not in ['exr', 'dpx', 'mov', 'jpg', 'png', 'tiff', 'tif', 'ma', 'mb', 'cube', 'csp']:
            # move these files to the extra_files dir
            extra_destfile = os.path.join(extra_files_dir, filename)
            if os.path.exists(extra_destfile):
                os.unlink(extra_destfile)
            shutil.move(filepath, extra_files_dir)
            return
        # deal with stills
        else:
            if fileext in ['jpg', 'png']:
                version_name_match = version_name_wframe_regexp.search(filename)
                if version_name_match:
                    filename_destination = '%s_%s_frame%s_%s.%s'%(version_name_match.group(1), version_name_match.group(2), version_name_match.group(4), version_name_match.group(3), fileext)
                    print('Info: %s will be renamed to %s.' % (filename, filename_destination))
                    os.rename(os.path.join(filedir, filename), os.path.join(filedir, filename_destination))
                    b_renamed = True
            elif fileext in ['cube', 'csp']:
                lut_destfile = os.path.join(lut_files_dir, filename)
                if not lut_destfile == filepath:
                    if os.path.exists(lut_destfile):
                        os.unlink(lut_destfile)
                    shutil.move(filepath, lut_files_dir)
                else:
                    print('Info: LUT file is already in the luts directory.')

    filebase_new = filebase
    for filename_extra_regexp in filename_extras_regexp_list:
        filename_extra_match = filename_extra_regexp.search(filebase)
        if filename_extra_match:
            filebase_new = filename_extra_match.group(1)
            new_filename_array = filename_array
            new_filename_array[0] = filebase_new
            filename_destination = '.'.join(new_filename_array)
            if not b_renamed:
                print('Info: %s will be renamed to %s.'%(filename, filename_destination))
                os.rename(os.path.join(filedir, filename), os.path.join(filedir, filename_destination))

    # account for frame numbers in version names of STILLS only
    if fileext in ['jpg', 'png']:
        if len(filename_array) > 2:
            filebase_new = "%s.%s"%(filebase_new, filename_array[1])
    master_files_list.append(filebase_new)
    try:
        master_files_ext_dict[filebase_new].append(fileext)
    except KeyError:
        master_files_ext_dict[filebase_new] = []
        master_files_ext_dict[filebase_new].append(fileext)


# make a "extra_files" folder if one does not exist
if not os.path.exists(extra_files_dir):
    os.makedirs(extra_files_dir)
else:
    if not os.path.isdir(extra_files_dir):
        raise RuntimeError("Path %s is not a directory!"%extra_files_dir)

# make a "lut_files" folder if one does not exist
if not os.path.exists(lut_files_dir):
    os.makedirs(lut_files_dir)
else:
    if not os.path.isdir(lut_files_dir):
        raise RuntimeError("Path %s is not a directory!"%lut_files_dir)

# make sure the directories are named correctly
directory_regexp_text_list = [r'(avid)', r'(vfx)', r'(exr)', r'(support_files)', r'(luts)', r'(matte)', r'(jpg)', r'(nuke)', r'(maya)', r'(tif)']
directory_regexp_list = [re.compile(pattern) for pattern in directory_regexp_text_list]

directory_list = os.listdir(dirpath)
for file in directory_list:
    if os.path.isdir(os.path.join(dirpath, file)):
        for directory_regexp in directory_regexp_list:
            directory_regexp_match = directory_regexp.search(file)
            if directory_regexp_match:
                match_group = directory_regexp_match.group(1)
                if file != match_group:
                    print('Info: directory %s will be renamed to %s.'%(file, match_group))
                    os.rename(os.path.join(dirpath, file), os.path.join(dirpath, match_group))

for dirname, subdirlist, filelist in os.walk(dirpath):
    for fname in filelist:
        check_file_naming(os.path.join(dirname, fname))

uniq_master_files_list = sorted(set(master_files_list))
master_files_dict = {}
for file in uniq_master_files_list:
    tmp_extension_list = []
    try:
        tmp_extension_list = master_files_ext_dict[file]
    except KeyError:
        pass

    tmp_file_dict = {'type' : None,
                     'link' : None,
                     'task' : None,
                     'subreason' : None,
                     'desc' : None,
                     'first' : None,
                     'last' : None,
                     'count' : None,
                     'extensions' : tmp_extension_list}
    print('Info: Found Version: %s'%file)
    master_files_dict[file] = tmp_file_dict
    version_match = version_regexp.search(file)
    if version_match:
        version_number = int(version_match.group(1))
        if version_number == 0:
            master_files_dict[file]['subreason'] = 'Scan Check'
        else:
            master_files_dict[file]['subreason'] = 'Final Comp'
    for tmp_ext in tmp_file_dict['extensions']:
        if tmp_ext in ['png', 'jpg']:
            master_files_dict[file]['subreason'] = 'Concept'
    shot_match = shot_regexp.search(file)
    if shot_match:
        master_files_dict[file]['type'] = 'Shot'
        master_files_dict[file]['link'] = shot_match.group(1)
        master_files_dict[file]['task'] = 'Final'
    else:
        sequence_match = sequence_regexp.search(file)
        if sequence_match:
            master_files_dict[file]['type'] = 'Sequence'
            master_files_dict[file]['link'] = sequence_match.group(1)
            master_files_dict[file]['task'] = 'R&D'
        else:
            print(
                'Warning: Unable to determine entity type for Version %s. Setting to Asset, but this will need to be manually adjusted.' % file)
            master_files_dict[file]['type'] = 'Asset'
            master_files_dict[file]['link'] = 'Miscellaneous'
            master_files_dict[file]['task'] = 'R&D'

    b_subform_match = False
    for subform_row in subform_rows:
        tmp_subform_version_code = ''
        tmp_file = file
        if tmp_file.find('.') != -1:
            tmp_file = tmp_file.split('.')[0]
        try:
            tmp_subform_version_code = subform_row['Shot/Asset']
        except KeyError:
            print('Warning: Submission form does not have Shot/Asset column. Will try with Version Name instead.')
            try:
                tmp_subform_version_code = subform_row['Version Name']
            except KeyError:
                print('Warning: Submission form does not have Version Name column either. Skipping.')
                continue

        frame_hack_removal = version_name_hack_regexp.sub('', file)
        new_file = file
        if frame_hack_removal != file:
            new_file = frame_hack_removal
        if tmp_subform_version_code.lower().find(new_file.lower()) != -1:
            b_subform_match = True
            print('Info: Found record for %s in submission form.'%file)
            try:
                master_files_dict[file]['desc'] = subform_row['Submission Notes']
                master_files_dict[file]['first'] = subform_row['Comp Start Frame']
                master_files_dict[file]['last'] = subform_row['Comp End Frame']
                master_files_dict[file]['count'] = subform_row['Duration']
            except KeyError:
                print('Warning: This submission form does not have columns for one or more of the following: Submission Notes, Comp Start Frame, Comp End Frame, or Duration.')
            break

    if not b_subform_match:
        master_files_dict[file]['desc'] = 'Shotgun Toolkit Create Delivery'

new_master_files_dict = {}
for tmp_version_name in master_files_dict.keys():
    new_version_name = tmp_version_name
    version_name_match = version_name_wframe_regexp.search(new_version_name)
    if version_name_match:
        new_version_name = '%s_%s_frame%s_%s'%(version_name_match.group(1), version_name_match.group(2), version_name_match.group(4), version_name_match.group(3))
    new_master_files_dict[new_version_name] = master_files_dict[tmp_version_name]

master_files_dict = new_master_files_dict
dir_basename = os.path.basename(dirpath)

csv_filepath = os.path.join(dirpath, '%s.csv'%dir_basename)
print('Info: Writing out CSV: %s'%csv_filepath)

with open(csv_filepath, 'w') as csvfile:
    writer = csv.DictWriter(csvfile, fieldnames=headers)
    writer.writeheader()

    rowdict = None
    for version_code in sorted(master_files_dict.keys()):
        print('Adding version %s to CSV file...'%version_code)
        rowdict = {}
        rowdict['Version Name'] = version_code
        rowdict['Link'] = master_files_dict[version_code]['link']
        rowdict['Task'] = master_files_dict[version_code]['task']
        rowdict['Type'] = 'Comp'
        rowdict['Submitted For'] = master_files_dict[version_code]['subreason']
        rowdict['Description'] = master_files_dict[version_code]['desc']
        rowdict['First Frame Text'] = master_files_dict[version_code]['first']
        rowdict['Last Frame Text'] = master_files_dict[version_code]['last']
        rowdict['Frame Count Text'] = master_files_dict[version_code]['count']
        writer.writerow(rowdict)

print('Successfully wrote out CSV file %s.'%csv_filepath)

# app = QtWidgets.QApplication([])
# msg = QtWidgets.QMessageBox()
# msg.setIcon(QtWidgets.QMessageBox.Information)
# msg.setText('Successfully completed Vendor Delivery Translation.\nCSV File: \n%s'%csv_filepath)
# msg.setWindowTitle('Delivery Prep')
# app.exec_()
# returnval = msg.show()
# app.quit()





