#!/usr/local/bin/python

import sys
from openpyxl import load_workbook
import os
import re
import ConfigParser
import db_access as DB
import getpass

supported_filetypes = None
version_name_transforms = {}
shot_name_ch = None
version_name_ch = None
note_body_ch = None
l_notes = []
ihdb = None
show_code = None

spreadsheet_file = ""
config = None

def usage(message):
    global supported_filetypes
    print """
Error: %s

This Python script is designed to take a spreadsheet containing notes and import them into a database.

Usage: note_ingest.py /path/to/spreadsheet.xlsx

Supported file types: %s
"""%(message, ', '.join(supported_filetypes))
    return

def shot_final(dbversion):
    global ihdb, config, show_code
    shot_status_final = config.get(show_code, 'shot_status_final')
    task_status_final = config.get(show_code, 'task_status_final')
    version_status_final = config.get(show_code, 'version_status_final')
    print "INFO: Version %s is final."%dbversion.g_version_code
    print "INFO: Setting shot status = %s, task status = %s, and version status = %s."%(shot_status_final, task_status_final, version_status_final)
    dbversion.g_shot.g_status = shot_status_final
    dbversion.g_task.g_status = task_status_final
    dbversion.g_status = version_status_final
    ihdb.update_shot_status(dbversion.g_shot)
    ihdb.update_task_status(dbversion.g_task)
    ihdb.update_version_status(dbversion)
    
def shot_pending_2k(dbversion):
    global ihdb, config, show_code
    shot_status_p2k = config.get(show_code, 'shot_status_p2k')
    task_status_p2k = config.get(show_code, 'task_status_p2k')
    version_status_p2k = config.get(show_code, 'version_status_p2k')
    print "INFO: Version %s is pending 2K approval."%dbversion.g_version_code
    print "INFO: Setting shot status = %s, task status = %s, and version status = %s."%(shot_status_p2k, task_status_p2k, version_status_p2k)
    dbversion.g_shot.g_status = shot_status_p2k
    dbversion.g_task.g_status = task_status_p2k
    dbversion.g_status = version_status_p2k
    ihdb.update_shot_status(dbversion.g_shot)
    ihdb.update_task_status(dbversion.g_task)
    ihdb.update_version_status(dbversion)

def shot_cbb(dbversion):
    global ihdb, config, show_code
    shot_status_cbb = config.get(show_code, 'shot_status_cbb')
    task_status_cbb = config.get(show_code, 'task_status_cbb')
    version_status_cbb = config.get(show_code, 'version_status_cbb')
    print "INFO: Version %s is CBB."%dbversion.g_version_code
    print "INFO: Setting shot status = %s, task status = %s, and version status = %s."%(shot_status_cbb, task_status_cbb, version_status_cbb)
    dbversion.g_shot.g_status = shot_status_cbb
    dbversion.g_task.g_status = task_status_cbb
    dbversion.g_status = version_status_cbb
    ihdb.update_shot_status(dbversion.g_shot)
    ihdb.update_task_status(dbversion.g_task)
    ihdb.update_version_status(dbversion)

def shot_temp_approved(dbversion):
    global ihdb, config, show_code
    shot_status_tmp = config.get(show_code, 'shot_status_tmp')
    task_status_tmp = config.get(show_code, 'task_status_tmp')
    version_status_tmp = config.get(show_code, 'version_status_tmp')
    print "INFO: Version %s is approved for temp."%dbversion.g_version_code
    print "INFO: Setting shot status = %s, task status = %s, and version status = %s."%(shot_status_tmp, task_status_tmp, version_status_tmp)
    dbversion.g_shot.g_status = shot_status_tmp
    dbversion.g_task.g_status = task_status_tmp
    dbversion.g_status = version_status_tmp
    ihdb.update_shot_status(dbversion.g_shot)
    ihdb.update_task_status(dbversion.g_task)
    ihdb.update_version_status(dbversion)
    
def shot_notes(dbversion):
    global ihdb, config, show_code
    shot_status_notes = config.get(show_code, 'shot_status_notes')
    task_status_notes = config.get(show_code, 'task_status_notes')
    version_status_notes = config.get(show_code, 'version_status_notes')
    print "INFO: Version %s has outstanding notes."%dbversion.g_version_code
    print "INFO: Setting shot status = %s, task status = %s, and version status = %s."%(shot_status_notes, task_status_notes, version_status_notes)
    dbversion.g_shot.g_status = shot_status_notes
    dbversion.g_task.g_status = task_status_notes
    dbversion.g_status = version_status_notes
    ihdb.update_shot_status(dbversion.g_shot)
    ihdb.update_task_status(dbversion.g_task)
    ihdb.update_version_status(dbversion)
    
shot_triggers = { 'shot_final' : shot_final, 
                  'shot_pending_2k' : shot_pending_2k,
                  'temp_approved' : shot_temp_approved,
                  'shot_cbb' : shot_cbb }

shot_triggers_keywords = {}

def load_config():
    global config, show_code
    if config == None:
        config = ConfigParser.ConfigParser()
        show_config_path = None
        try:
            show_config_path = os.environ['IH_SHOW_CFG_PATH']
            show_code = os.environ['IH_SHOW_CODE']
        except KeyError:
            raise RuntimeError("This system does not have an IH_SHOW_CFG_PATH environment variable defined.")
        if not os.path.exists(show_config_path):
            raise RuntimeError("The IH_SHOW_CFG_PATH environment variable is defined on this system with value %s, but no file exists at that location."%show_config_path)
        try:
            config.read(show_config_path)
        except:
            raise
 
def open_spreadsheet():
    global spreadsheet_file
    global version_name_ch
    global shot_name_ch
    global note_body_ch
    global l_notes
    global version_name_transforms
    
    wb = load_workbook(spreadsheet_file)
    ws = wb[wb.sheetnames[0]]
    colnames = [shot_name_ch, version_name_ch, note_body_ch]
    col_indices = {n: cell.value for n, cell in enumerate(ws.rows.next()) 
               if cell.value in colnames}
    
    row_num = 1
    for row in ws.iter_rows(min_row=2):
        row_num = row_num + 1
        d_tmp_note = {}
        b_hasblanks = False
        for index, cell in enumerate(row):
            if index in col_indices:
                if cell.value == None or len(cell.value) < 1:
                    print "ERROR in row %d: column %d is blank."%(row_num, index)
                    b_hasblanks = True
        if b_hasblanks:
            continue
        for index, cell in enumerate(row):
            if index in col_indices:
                if col_indices[index] == shot_name_ch:
                    d_tmp_note['shot_name'] = cell.value
                elif col_indices[index] == version_name_ch:
                    version_name = cell.value
                    for strxf in version_name_transforms.keys():
                        version_name = version_name.replace(strxf, version_name_transforms[strxf])
                    d_tmp_note['version_name'] = version_name
                elif col_indices[index] == note_body_ch:
                    d_tmp_note['note_body'] = cell.value
        l_notes.append(d_tmp_note)
    
def insert_notes():
    global l_notes
    global spreadsheet_file
    global ihdb
    global shot_triggers_keywords
    note_keyword_match = False
    ihdb = DB.DBAccessGlobals.get_db_access()
    notes_from = ihdb.fetch_artist_from_username(getpass.getuser())
    t_subject = os.path.splitext(os.path.basename(spreadsheet_file))[0]
    if not notes_from:
        print "ERROR: Unable to get user object from database for username %s. Program will exit."%getpass.getuser()
        return
    for d_note in l_notes:
        t_shot = d_note['shot_name']
        t_version_name = d_note['version_name']
        t_note_body = d_note['note_body']
        if not t_version_name:
            print "ERROR: Version name is blank. Skipping."
            continue
        if not t_shot:
            print "ERROR: Shot name is blank. Skipping."
            continue
        if not t_note_body:
            print "ERROR: Note for version %s is blank. Skipping."%t_version_name
            continue
        dbshot = ihdb.fetch_shot(t_shot)
        if not dbshot:
            print "ERROR: Unable to retrieve shot object from database for %s."%t_shot
            continue
        # dbversion = ihdb.fetch_version(t_version_name, dbshot)
        dbversion = None
        dbversions = ihdb.fetch_versions_for_shot(dbshot)
        for tmp_version in dbversions:
            tmp_version_match = re.search(t_version_name, tmp_version.g_version_code)
            if tmp_version_match:
                print "INFO: Found version match: %s"%tmp_version.g_version_code
                t_version_name = tmp_version.g_version_code
                dbversion = ihdb.fetch_version_from_id(tmp_version.g_dbid)
                break
        if not dbversion:
            print "ERROR: Unable to retrieve version object from database for %s."%t_version_name
            continue
        dbnotes = ihdb.fetch_notes_for_version(dbversion)
        b_notes_dupe = False
        for dbnote in dbnotes:
            if dbnote.g_body == t_note_body:
                print "WARNING: A note with the same content already exists in the database for version %s. Skipping."%t_version_name
                b_notes_dupe = True
                break
        if not b_notes_dupe:
            print "INFO: Creating new note in database for version %s."%t_version_name
            new_note = DB.Note(t_subject, dbversion.g_artist, notes_from, [dbshot, dbversion], t_note_body, config.get('note_ingest', 'default_note_type'), -1)
            ihdb.create_note(new_note)
            print "INFO: Note successfully created with database ID = %d."%new_note.g_dbid
        note_keyword_match = False
        for trigger in shot_triggers_keywords.keys():
            if trigger in t_note_body.lower():
                note_keyword_match = True
                print "INFO: Found trigger keyword %s in note body. Executing function %s(dbshot)."%(t_note_body, shot_triggers_keywords[trigger].__name__)
                shot_triggers_keywords[trigger](dbversion)
                break
        if not note_keyword_match:
            shot_notes(dbversion)
            

# make sure that there is one command line argument, which is the path to a valid spreadsheet
if __name__ == "__main__":

    if len(sys.argv) < 2:
        usage("Please provide a path to a valid spreadsheet file as the first command line argument.")
        exit()
    spreadsheet_file = sys.argv[1]
    if not os.path.exists(spreadsheet_file):
        usage("The path to the file provided on the command line does not exist.")
        exit()
    spreadsheet_ext = os.path.splitext(spreadsheet_file)[1]

    try:
        load_config()
        supported_filetypes = config.get('note_ingest', 'supported_filetypes').split(',')
        for xf in config.get('note_ingest', 'version_name_transforms').split(','):
            version_name_transforms[xf.split('|')[0]] = xf.split('|')[1]
        shot_name_ch = config.get('note_ingest', 'shot_name')
        version_name_ch = config.get('note_ingest', 'version_name')
        note_body_ch = config.get('note_ingest', 'note_body')
        note_triggers_txt = config.get('note_ingest', 'note_body_triggers')
        for tmp_trigger in note_triggers_txt.split(','):
            tmp_keyword, tmp_func_name = tmp_trigger.split('|')
            shot_triggers_keywords[tmp_keyword] = shot_triggers[tmp_func_name]
    except:
        usage(sys.exc_info()[1])
        exit()
        
    if not spreadsheet_ext in supported_filetypes:
        usage("Spreadsheet file type provided on command line is not supported.")
        exit()

    open_spreadsheet()
    insert_notes()
        
